#!/usr/bin/env python3
"""
Executive Email Scraper for Non-Profit Sponsorship Outreach
Uses Hunter.io API to find executive emails from company domains
"""

import json
import requests
import csv
import os
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from dotenv import load_dotenv
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

HUNTER_API_BASE = "https://api.hunter.io/v2"


def load_hunter_api_keys(env_path: Optional[Path] = None) -> List[str]:
    """Load all Hunter API keys from .env for automatic rotation on rate limit.
    Supports both uncommented (HUNTER_API_KEY=...) and commented (# HUNTER_API_KEY=...) lines.
    Only matches exact env-var style lines for commented keys (avoids false positives in freeform comments).
    """
    keys = []
    path = env_path or Path(__file__).resolve().parent / ".env"
    if path.exists():
        with open(path, encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if s.startswith("HUNTER_API_KEY="):
                    key = s.split("=", 1)[1].strip()
                    if key:
                        keys.append(key)
                elif re.match(r"^\s*#\s*HUNTER_API_KEY=(.+)$", s):
                    # Only match lines that are purely a commented env var (not "Use HUNTER_API_KEY=..." etc.)
                    key = s.split("HUNTER_API_KEY=", 1)[1].strip()
                    if key:
                        keys.append(key)
    if not keys:
        single = os.getenv("HUNTER_API_KEY")
        if single:
            keys.append(single)
    return keys

# Checkpoint: save progress every N companies so resume after rate limit is possible
CHECKPOINT_FILE = os.getenv("CHECKPOINT_FILE", "scraper_checkpoint.json")
CHECKPOINT_SAVE_EVERY = max(1, int(os.getenv("CHECKPOINT_SAVE_EVERY", "10")))

# Rate limiting configuration
RATE_LIMIT_DELAY = float(os.getenv("RATE_LIMIT_DELAY", "2"))
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))
RETRY_DELAY = int(os.getenv("RETRY_DELAY", "5"))


def load_checkpoint(path: str = CHECKPOINT_FILE) -> Tuple[List[Dict], List[Dict]]:
    """Load results and no_results from checkpoint file. Returns ([], []) if missing or invalid."""
    if not os.path.exists(path):
        return [], []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return [], []
        results = data.get("results", [])
        no_results = data.get("no_results", [])
        if not isinstance(results, list):
            results = []
        if not isinstance(no_results, list):
            no_results = []
        # Ensure we only keep dict-like rows (safety for corrupted checkpoint)
        results = [r for r in results if isinstance(r, dict)]
        no_results = [r for r in no_results if isinstance(r, dict)]
        return (results, no_results)
    except (json.JSONDecodeError, IOError, OSError) as e:
        print(f"Warning: Could not load checkpoint {path}: {e}")
        return [], []


def save_checkpoint(
    results: List[Dict],
    no_results: List[Dict],
    path: str = CHECKPOINT_FILE,
) -> None:
    """Write results and no_results to checkpoint file (atomic write: temp then rename)."""
    try:
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"results": results, "no_results": no_results}, f, indent=2)
        os.replace(tmp, path)  # atomic on POSIX and Windows
    except IOError as e:
        print(f"Warning: Could not save checkpoint {path}: {e}")


def extract_domain(url_or_domain: str) -> str:
    """
    Extract clean domain name from URL, @domain format, or plain domain.

    Supports:
    - URLs: https://www.example.com/, http://example.com/path
    - @domain: @example.com
    - Plain domain: example.com

    Returns:
        Clean domain name (e.g., 'example.com')
    """
    if not url_or_domain:
        return ""

    # Remove whitespace
    url_or_domain = url_or_domain.strip()

    if not url_or_domain:
        return ""

    # Strip leading @ if present (e.g. @smuckers.com -> smuckers.com)
    if url_or_domain.startswith("@"):
        url_or_domain = url_or_domain[1:].strip()
        if not url_or_domain:
            return ""

    # If it's already a clean domain (no protocol), return as-is
    if not url_or_domain.startswith(("http://", "https://")):
        # Remove any trailing slashes or paths
        domain = url_or_domain.split("/")[0]
        return domain.lower()

    # Parse URL to extract domain
    try:
        parsed = urlparse(url_or_domain)
        domain = parsed.netloc.lower()

        if not domain:
            return ""

        # Remove 'www.' prefix if present
        if domain.startswith("www."):
            domain = domain[4:]

        return domain
    except Exception as e:
        print(f"Warning: Could not parse URL '{url_or_domain}': {e}")
        # Fallback: try to extract domain using regex
        domain_match = re.search(r"([a-zA-Z0-9-]+\.[a-zA-Z]{2,})", url_or_domain)
        if domain_match:
            return domain_match.group(1).lower()
        return ""


def clean_domains(domains: List[str]) -> List[str]:
    """
    Clean and deduplicate a list of domains/URLs

    Args:
        domains: List of URLs or domain names

    Returns:
        List of clean, unique domain names
    """
    cleaned_domains = []
    seen = set()

    for domain in domains:
        clean_domain = extract_domain(domain)
        if clean_domain and clean_domain not in seen:
            cleaned_domains.append(clean_domain)
            seen.add(clean_domain)

    return cleaned_domains


def _result_rows_for_domain(
    domain: str,
    search_result: Dict,
    bp_member: str,
    parse_date: str,
    executives_only: bool = True,
) -> Tuple[List[Dict], Optional[Dict]]:
    """
    Convert one domain's search_domain() result into rows for results and no_results.
    Returns (list of result row dicts, no_result row dict or None).
    """
    if search_result.get("emails"):
        emails = search_result["emails"]
        if executives_only:
            emails = [e for e in emails if e.get("is_executive")]
        if emails:
            rows = []
            for email in emails:
                rows.append(
                    {
                        "Domain": domain,
                        "Company": search_result.get("company", domain),
                        "Email": email.get("email"),
                        "First Name": email.get("first_name"),
                        "Last Name": email.get("last_name"),
                        "Position": email.get("position"),
                        "Department": email.get("department"),
                        "Confidence": email.get("confidence"),
                        "BP Member": bp_member,
                        "Parse Date": parse_date,
                    }
                )
            return (rows, None)
        # Emails found but none are executives
        no_result = {
            "Domain": domain,
            "Company": search_result.get("company", domain),
            "BP Member": bp_member,
            "Reason": "No executive emails found",
            "Parse Date": parse_date,
        }
        return ([], no_result)
    # No emails found at all
    no_result = {
        "Domain": domain,
        "Company": search_result.get("company", domain),
        "BP Member": bp_member,
        "Reason": search_result.get("error", "No emails in database"),
        "Parse Date": parse_date,
    }
    return ([], no_result)


class EmailScraper:
    def __init__(self, api_keys: List[str]):
        if not api_keys:
            raise ValueError("At least one API key is required")
        api_keys = [k.strip() for k in api_keys if k and k.strip()]
        if not api_keys:
            raise ValueError("API key cannot be empty")
        self.api_keys = api_keys
        self.current_key_index = 0
        self.results = []
        self.no_results = []  # Track companies with no executive emails found
        self.last_request_time = 0  # Track last API request time for rate limiting

    @property
    def api_key(self) -> str:
        """Current API key for backward compatibility."""
        return self.api_keys[self.current_key_index]

    def _wait_for_rate_limit(self):
        """Implement rate limiting by ensuring minimum delay between requests"""
        current_time = time.time()
        time_since_last_request = current_time - self.last_request_time

        if time_since_last_request < RATE_LIMIT_DELAY:
            sleep_time = RATE_LIMIT_DELAY - time_since_last_request
            print(f"  Rate limiting: waiting {sleep_time:.1f}s...")
            time.sleep(sleep_time)

        self.last_request_time = time.time()

    def _make_api_request(self, url: str, params: dict, domain: str = "") -> Optional[Dict]:
        """
        Make API request with retry logic, exponential backoff, and automatic key rotation on 429.

        Args:
            url: API endpoint URL
            params: Request parameters (api_key is updated when rotating keys)
            domain: Domain being queried (for error messages)

        Returns:
            Response JSON data or None if all retries and keys failed
        """
        key_index = self.current_key_index
        while key_index < len(self.api_keys):
            params["api_key"] = self.api_keys[key_index]

            for attempt in range(MAX_RETRIES):
                try:
                    # Enforce rate limiting before each request
                    self._wait_for_rate_limit()

                    response = requests.get(url, params=params, timeout=10)

                    # Handle rate limiting (429 = usage limit, 403 = request rate limit per Hunter API docs)
                    if response.status_code in (403, 429):
                        retry_delay = RETRY_DELAY * (2 ** attempt)  # Exponential backoff
                        print(f"  Rate limit hit (HTTP {response.status_code}). Retrying in {retry_delay}s (attempt {attempt + 1}/{MAX_RETRIES})...")
                        time.sleep(retry_delay)
                        continue

                    # Handle other HTTP errors
                    response.raise_for_status()
                    self.current_key_index = key_index
                    try:
                        return response.json()
                    except json.JSONDecodeError as e:
                        print(f"Error: Invalid JSON response for {domain}: {e}")
                        return None

                except requests.exceptions.Timeout:
                    print(f"  Request timeout for {domain}. Retrying (attempt {attempt + 1}/{MAX_RETRIES})...")
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(RETRY_DELAY)
                        continue
                    else:
                        print(f"Error: Request timeout for {domain} after {MAX_RETRIES} attempts")
                        return None

                except requests.exceptions.HTTPError as e:
                    if e.response.status_code == 401:
                        print(f"Error: Invalid API key. Check your HUNTER_API_KEY in .env file.")
                        return None
                    elif e.response.status_code in (403, 429):
                        # Already handled above, but keeping for safety
                        continue
                    else:
                        print(f"Error: HTTP {e.response.status_code} for {domain}")
                        return None

                except requests.exceptions.RequestException as e:
                    print(f"  Network error for {domain}: {e}")
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(RETRY_DELAY)
                        continue
                    else:
                        print(f"Error: Failed to fetch {domain} after {MAX_RETRIES} attempts")
                        return None

            # All retries exhausted for this key - try next key if available
            if key_index < len(self.api_keys) - 1:
                key_index += 1
                self.current_key_index = key_index
                print(f"  Rate limit hit. Switching to backup API key {key_index + 1} of {len(self.api_keys)}...")
            else:
                print(f"Error: Rate limit exceeded for {domain}. Please wait or upgrade your Hunter.io plan.")
                return None

        return None

    def search_domain(self, domain: str, role: str = None) -> Dict:
        """
        Search for emails from a specific domain

        Args:
            domain: Company domain (e.g., 'stripe.com')
            role: Optional role filter (e.g., 'executive', 'ceo', 'cfo')
        """
        if not domain or not domain.strip():
            print("Warning: Empty domain provided")
            return {"domain": domain, "emails": []}

        url = f"{HUNTER_API_BASE}/domain-search"
        params = {
            "domain": domain,
            "api_key": self.api_key,
            "type": "personal",  # Focus on individual emails, not generic ones
        }

        if role:
            params["role"] = role

        # Use the new rate-limited API request method
        data = self._make_api_request(url, params, domain)

        if data and data.get("data"):
            return self._parse_results(data["data"], domain)
        elif data is None:
            # Request failed after retries
            return {"domain": domain, "emails": [], "error": "Request failed"}
        else:
            print(f"No results found for {domain}")
            return {"domain": domain, "emails": []}

    def find_email(
        self, domain: str, first_name: str, last_name: str
    ) -> Optional[Dict]:
        """
        Find a specific person's email

        Args:
            domain: Company domain
            first_name: Person's first name
            last_name: Person's last name
        """
        url = f"{HUNTER_API_BASE}/email-finder"
        params = {
            "domain": domain,
            "first_name": first_name,
            "last_name": last_name,
            "api_key": self.api_key,
        }

        # Use the new rate-limited API request method
        data = self._make_api_request(url, params, domain)

        if data and data.get("data") and data["data"].get("email"):
            return {
                "email": data["data"]["email"],
                "first_name": data["data"].get("first_name"),
                "last_name": data["data"].get("last_name"),
                "position": data["data"].get("position"),
                "confidence": data["data"].get("score"),
                "domain": domain,
            }
        return None

    def _parse_results(self, data: Dict, domain: str) -> Dict:
        """Parse Hunter.io API response"""
        emails = []

        for email_data in data.get("emails", []):
            # Filter for executive roles
            position = email_data.get("position") or ""
            position = position.lower() if position else ""
            executive_keywords = [
                "ceo",
                "cfo",
                "cto",
                "coo",
                "chief",
                "president",
                "founder",
                "director",
                "vp",
                "vice president",
                "head",
            ]

            is_executive = any(keyword in position for keyword in executive_keywords)

            email_info = {
                "email": email_data.get("value"),
                "first_name": email_data.get("first_name"),
                "last_name": email_data.get("last_name"),
                "position": email_data.get("position"),
                "department": email_data.get("department"),
                "confidence": email_data.get("confidence"),
                "is_executive": is_executive,
                "domain": domain,
            }
            emails.append(email_info)

        return {
            "domain": domain,
            "company": data.get("organization"),
            "emails": emails,
            "total_found": len(emails),
        }

    def scrape_companies(
        self,
        domains: List[str],
        domain_to_member: Dict[str, str] = None,
        executives_only: bool = True,
    ) -> List[Dict]:
        """
        Scrape multiple company domains

        Args:
            domains: List of company domains
            domain_to_member: Mapping of domains to BP member names
            executives_only: If True, filter for executive positions only
        """
        all_results = []
        no_results = []
        parse_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if domain_to_member is None:
            domain_to_member = {}

        for domain in domains:
            print(f"Searching {domain}...")
            result = self.search_domain(domain)
            bp_member = domain_to_member.get(domain, "Unknown")

            if result.get("emails"):
                emails = result["emails"]

                if executives_only:
                    emails = [e for e in emails if e.get("is_executive")]

                if emails:
                    for email in emails:
                        all_results.append(
                            {
                                "Domain": domain,
                                "Company": result.get("company", domain),
                                "Email": email.get("email"),
                                "First Name": email.get("first_name"),
                                "Last Name": email.get("last_name"),
                                "Position": email.get("position"),
                                "Department": email.get("department"),
                                "Confidence": email.get("confidence"),
                                "BP Member": bp_member,
                                "Parse Date": parse_date,
                            }
                        )
                    print(f"  Found {len(emails)} executive email(s)")
                else:
                    # Emails found but none are executives
                    no_results.append(
                        {
                            "Domain": domain,
                            "Company": result.get("company", domain),
                            "BP Member": bp_member,
                            "Reason": "No executive emails found",
                            "Parse Date": parse_date,
                        }
                    )
                    print(f"  No executive emails found")
            else:
                # No emails found at all
                no_results.append(
                    {
                        "Domain": domain,
                        "Company": result.get("company", domain),
                        "BP Member": bp_member,
                        "Reason": result.get("error", "No emails in database"),
                        "Parse Date": parse_date,
                    }
                )
                print(f"  No emails found")

        self.results = all_results
        self.no_results = no_results
        return all_results

    def export_to_csv(self, filename: str = "executive_emails.csv"):
        """Export results to CSV file, appending to existing file if present"""
        if not self.results:
            print("No results to export")
            return

        file_exists = os.path.exists(filename)

        # Define expected headers to ensure consistency
        expected_headers = [
            "Domain",
            "Company",
            "Email",
            "First Name",
            "Last Name",
            "Position",
            "Department",
            "Confidence",
            "BP Member",
            "Parse Date",
        ]

        # Load existing emails to avoid duplicates
        existing_emails = set()
        headers_exist = False

        if file_exists:
            try:
                with open(filename, "r", newline="", encoding="utf-8") as f:
                    first_line = f.readline().strip()
                    # Check if first line looks like headers (starts with expected column names)
                    first_line_lower = first_line.lower()
                    headers_exist = (
                        first_line_lower.startswith("domain,")
                        or "domain,company,email" in first_line_lower
                    )
                    f.seek(0)  # Reset file pointer

                    if headers_exist:
                        reader = csv.DictReader(f)
                        for row in reader:
                            email = row.get("Email", "")
                            if email:
                                existing_emails.add(email.lower())
                    else:
                        # File exists but no headers - read as plain CSV
                        reader = csv.reader(f)
                        for row in reader:
                            if (
                                len(row) > 2
                            ):  # Make sure we have at least an email column
                                existing_emails.add(
                                    row[2].lower() if len(row) > 2 else ""
                                )
            except (IOError, csv.Error) as e:
                print(f"Warning: Could not read existing file {filename}: {e}")
                print("Creating new file...")

        # Filter out duplicates and empty emails
        new_results = [
            r
            for r in self.results
            if r.get("Email") and r.get("Email", "").lower() not in existing_emails
        ]

        if not new_results:
            print(
                f"\n✓ No new emails to add (all {len(self.results)} already exist in {filename})"
            )
            return

        # Write to file (create new or append)
        try:
            if not file_exists or not headers_exist:
                # Create new file with headers or rewrite existing file with headers
                if file_exists and not headers_exist:
                    # Backup existing content before overwriting
                    try:
                        with open(filename, "r", encoding="utf-8") as old:
                            old_content = old.read()
                        with open(filename + ".backup", "w", encoding="utf-8") as backup:
                            backup.write(old_content)
                    except IOError:
                        pass  # Best-effort backup
                with open(filename, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=expected_headers)
                    writer.writeheader()
                    writer.writerows(new_results)
            else:
                # Append to existing file with headers
                with open(filename, "a", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=expected_headers)
                    writer.writerows(new_results)

            print(
                f"\n✓ Added {len(new_results)} new emails to {filename} ({len(self.results) - len(new_results)} duplicates skipped)"
            )
        except IOError as e:
            print(f"Error: Could not write to {filename}: {e}")
            print("Please check file permissions and disk space.")

    def export_to_excel(self, filename: str = "executive_emails.xlsx"):
        """Export results to Excel file with two sheets, appending to existing file if present"""
        file_exists = os.path.exists(filename)

        # Define headers for both sheets
        success_headers = [
            "Domain",
            "Company",
            "Email",
            "First Name",
            "Last Name",
            "Position",
            "Department",
            "Confidence",
            "BP Member",
            "Parse Date",
        ]

        no_results_headers = [
            "Domain",
            "Company",
            "BP Member",
            "Reason",
            "Parse Date",
        ]

        # Load existing data to avoid duplicates
        existing_emails = set()
        existing_no_result_domains = set()

        if file_exists:
            try:
                wb = load_workbook(filename)

                # Read existing emails from first sheet
                if "Executive Emails" in wb.sheetnames:
                    ws = wb["Executive Emails"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > 2 and row[2]:  # Check email column
                            existing_emails.add(str(row[2]).lower())

                # Read existing no-result domains from second sheet
                if "No Results" in wb.sheetnames:
                    ws_no = wb["No Results"]
                    for row in ws_no.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:  # Check domain column
                            existing_no_result_domains.add(str(row[0]).lower())

            except Exception as e:
                print(f"Warning: Could not read existing file {filename}: {e}")
                print("Creating new file...")
                file_exists = False

        # Filter out duplicates and empty emails
        new_results = [
            r
            for r in self.results
            if r.get("Email") and r.get("Email", "").lower() not in existing_emails
        ]

        new_no_results = [
            r
            for r in self.no_results
            if r.get("Domain")
            and r.get("Domain", "").lower() not in existing_no_result_domains
            and r.get("Domain", "").lower()
            not in existing_emails  # Don't add if we found emails
        ]

        # Check if we have anything to add
        if not new_results and not new_no_results:
            print(f"\n✓ No new data to add (all data already exists in {filename})")
            return

        try:
            if file_exists:
                # Load existing workbook
                wb = load_workbook(filename)

                # Get or create sheets
                if "Executive Emails" in wb.sheetnames:
                    ws = wb["Executive Emails"]
                else:
                    ws = wb.create_sheet("Executive Emails", 0)
                    self._style_header_row(ws, success_headers)
                    self._set_column_widths(
                        ws,
                        {
                            "A": 20,
                            "B": 25,
                            "C": 30,
                            "D": 15,
                            "E": 15,
                            "F": 30,
                            "G": 20,
                            "H": 12,
                            "I": 15,
                            "J": 20,
                        },
                    )

                if "No Results" in wb.sheetnames:
                    ws_no = wb["No Results"]
                else:
                    ws_no = wb.create_sheet("No Results")
                    self._style_header_row(ws_no, no_results_headers)
                    self._set_column_widths(
                        ws_no, {"A": 20, "B": 25, "C": 15, "D": 35, "E": 20}
                    )
            else:
                # Create new workbook with both sheets
                wb = Workbook()
                ws = wb.active
                ws.title = "Executive Emails"
                self._style_header_row(ws, success_headers)
                self._set_column_widths(
                    ws,
                    {
                        "A": 20,
                        "B": 25,
                        "C": 30,
                        "D": 15,
                        "E": 15,
                        "F": 30,
                        "G": 20,
                        "H": 12,
                        "I": 15,
                        "J": 20,
                    },
                )

                # Create No Results sheet
                ws_no = wb.create_sheet("No Results")
                self._style_header_row(ws_no, no_results_headers)
                self._set_column_widths(
                    ws_no, {"A": 20, "B": 25, "C": 15, "D": 35, "E": 20}
                )

            # Add new successful results
            for result in new_results:
                ws.append(
                    [
                        result.get("Domain", ""),
                        result.get("Company", ""),
                        result.get("Email", ""),
                        result.get("First Name", ""),
                        result.get("Last Name", ""),
                        result.get("Position", ""),
                        result.get("Department", ""),
                        result.get("Confidence", ""),
                        result.get("BP Member", ""),
                        result.get("Parse Date", ""),
                    ]
                )

            # Add new no-result entries
            for result in new_no_results:
                ws_no.append(
                    [
                        result.get("Domain", ""),
                        result.get("Company", ""),
                        result.get("BP Member", ""),
                        result.get("Reason", ""),
                        result.get("Parse Date", ""),
                    ]
                )

            # Save workbook
            wb.save(filename)

            # Print summary
            messages = []
            if new_results:
                messages.append(f"{len(new_results)} emails added")
            if new_no_results:
                messages.append(f"{len(new_no_results)} no-result companies added")

            duplicates_skipped = (len(self.results) - len(new_results)) + (
                len(self.no_results) - len(new_no_results)
            )

            print(
                f"\n✓ {', '.join(messages)} to {filename} ({duplicates_skipped} duplicates skipped)"
            )

        except Exception as e:
            print(f"Error: Could not write to {filename}: {e}")
            print("Please check file permissions and disk space.")

    def _style_header_row(self, ws, headers):
        """Apply styling to header row"""
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _set_column_widths(self, ws, widths):
        """Set column widths for worksheet"""
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def load_domains(filename: str = "companies.txt") -> tuple[List[str], Dict[str, str]]:
    """Load company domains from file and extract clean domain names with BP member mapping"""
    raw_domains = []
    domain_to_member = {}
    current_member = "Unknown"

    if not os.path.exists(filename):
        print(f"WARNING: {filename} not found. Creating example file...")
        try:
            with open(filename, "w") as f:
                f.write(
                    "## Example Member\nstripe.com\nsalesforce.com\n# Add more company domains below (one per line)\n"
                )
        except IOError as e:
            print(f"Error: Could not create {filename}: {e}")
            return [], {}

        return ["stripe.com", "salesforce.com"], {
            "stripe.com": "Example Member",
            "salesforce.com": "Example Member",
        }

    try:
        with open(filename, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                # Check for BP member section headers (## Name)
                if line.startswith("## "):
                    current_member = line[3:].strip()  # Remove "## " prefix
                    if not current_member:
                        current_member = "Unknown"
                    continue
                # Skip empty lines and comments
                if line and not line.startswith("#"):
                    raw_domains.append(line)
                    # Map the raw domain/URL to current member
                    clean_domain = extract_domain(line)
                    if clean_domain:  # Only add non-empty domains
                        domain_to_member[clean_domain] = current_member
    except IOError as e:
        print(f"Error: Could not read {filename}: {e}")
        return [], {}

    # Extract and clean domains
    cleaned_domains = clean_domains(raw_domains)

    if not cleaned_domains:
        print(f"Warning: No valid domains found in {filename}")
        return [], {}

    print(
        f"Loaded {len(raw_domains)} entries, extracted {len(cleaned_domains)} unique domains"
    )
    if len(raw_domains) != len(cleaned_domains):
        print("Note: Some duplicates were removed during domain extraction")

    return cleaned_domains, domain_to_member


def compute_domains_to_retry(
    domains: List[str],
    results: List[Dict],
    no_results: List[Dict],
) -> Tuple[List[str], List[Dict], int]:
    """
    Compute which domains need to be scraped (including retries for "Request failed").

    Domains with "Request failed" in no_results are retried; other no_results are skipped.
    Returns (domains_to_do, filtered_no_results, retry_count).
    """
    scraped_domains = set()
    for r in results:
        if r.get("Domain"):
            scraped_domains.add(r["Domain"])
    for r in no_results:
        if r.get("Domain"):
            if r.get("Reason") != "Request failed":
                scraped_domains.add(r["Domain"])

    domains_to_do = [d for d in domains if d not in scraped_domains]

    retry_count = 0
    if domains_to_do:
        retry_domains = set(domains_to_do)
        retry_count = sum(1 for r in no_results if r.get("Domain") in retry_domains)
        no_results = [r for r in no_results if r.get("Domain") not in retry_domains]

    return (domains_to_do, no_results, retry_count)


def main():
    """Main execution function. Uses checkpoint JSON to resume after rate limits; Excel is written only when all domains are scraped."""

    # Load API keys (supports multiple keys for automatic rotation on rate limit)
    api_keys = load_hunter_api_keys()
    if not api_keys:
        print("ERROR: HUNTER_API_KEY not found!")
        print("Please create a .env file with your Hunter.io API key:")
        print("HUNTER_API_KEY=your_api_key_here")
        return

    try:
        scraper = EmailScraper(api_keys)
    except ValueError as e:
        print(f"ERROR: {e}")
        return

    # Load domains from file
    domains, domain_to_member = load_domains("companies.txt")

    if not domains:
        print("ERROR: No domains found in companies.txt")
        return

    # Load checkpoint so we can skip already-scraped domains and resume after rate limit
    results, no_results = load_checkpoint()
    domains_to_do, no_results, retry_count = compute_domains_to_retry(
        domains, results, no_results
    )
    scraped_domains = set(domains) - set(domains_to_do)

    print("Executive Email Scraper for Non-Profit Sponsorship")
    print("=" * 50)
    if scraped_domains:
        print(f"Resuming: {len(scraped_domains)} already scraped, {len(domains_to_do)} remaining.")
    if retry_count:
        print(f"(Retrying {retry_count} domain(s) that failed previously.)")
    print(f"Searching {len(domains_to_do)} companies...\n")

    if not domains_to_do:
        # All done from a previous run; just build Excel from checkpoint
        scraper.results = results
        scraper.no_results = no_results
        if results or no_results:
            excel_path = "executive_emails.xlsx"
            if os.path.exists(excel_path):
                os.remove(excel_path)
            scraper.export_to_excel(excel_path)
            print("All domains already scraped. Excel file updated from checkpoint.")
        else:
            print("All domains already scraped. No data in checkpoint; no Excel written.")
        return

    parse_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, domain in enumerate(domains_to_do):
        print(f"Searching {domain}...")
        search_result = scraper.search_domain(domain)
        bp_member = domain_to_member.get(domain, "Unknown")
        new_rows, no_row = _result_rows_for_domain(
            domain, search_result, bp_member, parse_date, executives_only=True
        )
        results.extend(new_rows)
        if no_row:
            no_results.append(no_row)

        if new_rows:
            print(f"  Found {len(new_rows)} executive email(s)")
        else:
            if no_row and no_row.get("Reason") == "No executive emails found":
                print("  No executive emails found")
            else:
                print("  No emails found")

        # Save checkpoint every N companies so progress survives rate limit / Ctrl+C
        if (i + 1) % CHECKPOINT_SAVE_EVERY == 0:
            save_checkpoint(results, no_results)
            print(f"  Checkpoint saved ({len(results)} emails, {len(no_results)} no-results so far)")

    # Final checkpoint save
    save_checkpoint(results, no_results)

    # Build Excel only when all domains have been scraped (from checkpoint as single source of truth)
    scraped_now = {r["Domain"] for r in results if r.get("Domain")} | {
        r["Domain"] for r in no_results if r.get("Domain")
    }
    if scraped_now >= set(domains):
        scraper.results = results
        scraper.no_results = no_results
        excel_path = "executive_emails.xlsx"
        if os.path.exists(excel_path):
            os.remove(excel_path)  # Write fresh from checkpoint, not append
        scraper.export_to_excel(excel_path)
        print(f"\n✓ All {len(domains)} domains scraped. Excel saved to {excel_path}")
    else:
        print(
            f"\nScraping paused (e.g. rate limit). Progress saved to {CHECKPOINT_FILE}. "
            "Change API key if needed and run again to resume."
        )


if __name__ == "__main__":
    main()
